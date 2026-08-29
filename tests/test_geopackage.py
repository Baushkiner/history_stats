"""Чтение GeoPackage без внешних зависимостей.

Образец собирается здесь же: файл SQLite с обязательной таблицей
`gpkg_geometry_columns` и полем геометрии в формате «заголовок GP + WKB».
Проверяется то, на чём разбор обычно ломается: необязательная рамка в
заголовке, вложенный порядок байт в WKB и мультиполигон из нескольких колец.
"""

import sqlite3
import struct
from pathlib import Path

import pytest

from histctx.sources.geopackage import (
    GeoPackageError, layers, parse_geometry, read_features,
)

ROOT = Path(__file__).resolve().parents[1]

SQUARE = [[0.0, 0.0], [0.0, 10.0], [10.0, 10.0], [10.0, 0.0], [0.0, 0.0]]
HOLE = [[2.0, 2.0], [4.0, 2.0], [4.0, 4.0], [2.0, 4.0], [2.0, 2.0]]


def wkb_polygon(rings: list[list]) -> bytes:
    out = struct.pack("<BII", 1, 3, len(rings))
    for ring in rings:
        out += struct.pack("<I", len(ring))
        out += struct.pack(f"<{2 * len(ring)}d", *[c for point in ring for c in point])
    return out


def wkb_multipolygon(polygons: list[list]) -> bytes:
    out = struct.pack("<BII", 1, 6, len(polygons))
    for rings in polygons:
        out += wkb_polygon(rings)
    return out


def gpkg_blob(wkb: bytes, *, envelope: int = 0) -> bytes:
    """Поле геометрии: сигнатура GP, версия, флаги, SRS и необязательная рамка."""
    flags = 0x01 | (envelope << 1)
    header = b"GP" + bytes([0, flags]) + struct.pack("<i", 4326)
    sizes = {0: 0, 1: 4, 2: 6, 3: 6, 4: 8}
    header += struct.pack(f"<{sizes[envelope]}d", *([0.0] * sizes[envelope]))
    return header + wkb


def build_gpkg(path: Path, rows: list[tuple]) -> Path:
    con = sqlite3.connect(path)
    con.execute("create table gpkg_geometry_columns "
                "(table_name text, column_name text, geometry_type_name text, srs_id int)")
    con.execute("insert into gpkg_geometry_columns values "
                "('districts_1897', 'geom', 'MULTIPOLYGON', 4326)")
    con.execute('create table districts_1897 (fid integer primary key, geom blob, '
                '"Name_RU" text, "prov_RU" text)')
    con.executemany('insert into districts_1897 (geom, "Name_RU", "prov_RU") '
                    "values (?, ?, ?)", rows)
    con.commit()
    con.close()
    return path


@pytest.fixture()
def sample(tmp_path) -> Path:
    return build_gpkg(tmp_path / "sample.gpkg", [
        (gpkg_blob(wkb_multipolygon([[SQUARE, HOLE]])), "Кемский", "Архангельская губерния"),
        (gpkg_blob(wkb_polygon([SQUARE]), envelope=1), "Онежский", "Архангельская губерния"),
    ])


def test_layer_is_found_by_its_geometry_column(sample):
    assert layers(sample) == [("districts_1897", "geom")]


def test_features_are_read_with_attributes(sample):
    feats = read_features(sample)
    assert [f["properties"]["Name_RU"] for f in feats] == ["Кемский", "Онежский"]
    assert feats[0]["properties"]["prov_RU"] == "Архангельская губерния"
    # `fid` — счётчик строк файла, в свойства он не идёт.
    assert "fid" not in feats[0]["properties"]


def test_multipolygon_keeps_its_hole(sample):
    geometry = read_features(sample)[0]["geometry"]
    assert geometry["type"] == "MultiPolygon"
    assert len(geometry["coordinates"]) == 1
    assert len(geometry["coordinates"][0]) == 2
    assert geometry["coordinates"][0][0][2] == [10.0, 10.0]


def test_envelope_in_the_header_is_skipped(sample):
    """Рамка в заголовке необязательна: пропустить её надо ровно по размеру."""
    geometry = read_features(sample)[1]["geometry"]
    assert geometry == {"type": "Polygon", "coordinates": [SQUARE]}


def test_geometry_is_read_in_big_endian_too():
    wkb = struct.pack(">BII", 0, 3, 1) + struct.pack(">I", 4)
    ring = [[1.0, 2.0], [3.0, 4.0], [5.0, 6.0], [1.0, 2.0]]
    wkb += struct.pack(">8d", *[c for point in ring for c in point])
    assert parse_geometry(gpkg_blob(wkb)) == {"type": "Polygon", "coordinates": [ring]}


def test_foreign_format_stops_the_read(tmp_path):
    with pytest.raises(GeoPackageError, match="GP"):
        parse_geometry(b"NOTAGPKGBLOB" * 4)
    with pytest.raises(GeoPackageError, match="WKB"):
        parse_geometry(gpkg_blob(struct.pack("<BI", 1, 1) + struct.pack("<dd", 1.0, 2.0)))
    plain = tmp_path / "plain.sqlite"
    sqlite3.connect(plain).execute("create table t (a int)")
    with pytest.raises(GeoPackageError):
        layers(plain)


def test_empty_geometry_is_not_an_error():
    """Пустая геометрия помечается флагом в заголовке — это не поломка файла."""
    flags = 0x01 | (1 << 4)
    blob = b"GP" + bytes([0, flags]) + struct.pack("<i", 4326)
    assert parse_geometry(blob) is None


def test_xlsx_sheet_title_survives_a_colon(tmp_path):
    """Регрессия: двоеточие в названии слоя роняло сбор после записи GeoJSON.

    openpyxl запрещает в имени листа `[]:*?/\\`. Очистка в проекте была, но
    ею пользовалась только многолистовая выгрузка; однолистовая ставила имя
    как есть. Слой «Населённые места: год основания» собирался, писал
    GeoJSON и падал на XLSX — команда возвращала ошибку при собранных данных.
    """
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
    from histctx.io_formats import write_xlsx
    from histctx.schema import LayerSpec

    spec = LayerSpec(slug="x", title="Населённые места: год основания",
                     group="admin", source="Викиданные", license="CC0")
    rec = spec.new_record(title="Псков", lat=57.8, lon=28.3, year_from=903)
    path = tmp_path / "x.xlsx"
    assert write_xlsx([rec], path, sheet_name=spec.title) == 1

    from openpyxl import load_workbook
    assert ":" not in load_workbook(path).sheetnames[0]
