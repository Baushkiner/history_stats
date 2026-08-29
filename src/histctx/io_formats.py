"""Выгрузка слоёв в GeoJSON и XLSX."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Sequence

from .schema import COLUMNS, COLUMNS_RU, ContextRecord


# Свойства, одинаковые у всех записей слоя: название слоя, источник, права.
# У слоя в тысячу точек они занимают место, у слоя в семнадцать тысяч —
# больше половины файла.
HOISTABLE = ("layer", "layer_title", "group", "source", "license", "confidence",
             "date_precision", "date_approx")


def write_geojson(records: Iterable[ContextRecord], path: Path, *, layer_title: str = "",
                  hoist_shared: bool = False) -> int:
    """Пишет FeatureCollection. Записи без координат пропускаются.

    `hoist_shared` выносит свойства, одинаковые у всех записей, на уровень
    коллекции — в поле `layer`. Для больших слоёв это разница в разы, но
    читателю файла нужно знать про это поле, поэтому по умолчанию выключено:
    молча менять формат уже отданных слоёв нельзя.
    """
    feats = [r.to_feature() for r in records if r.has_point]
    shared = _shared_properties(feats) if hoist_shared else {}
    for feat in feats:
        for key in shared:
            feat["properties"].pop(key, None)

    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "type": "FeatureCollection",
        "name": layer_title or path.stem,
        "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3/CRS84"}},
        "features": feats,
    }
    if shared:
        # Читается как «свойства слоя»: то же самое, что было бы у каждой
        # точки, вынесенное один раз.
        payload["layer"] = shared
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
    return len(feats)


def _shared_properties(feats: Sequence[dict]) -> dict:
    """Свойства из HOISTABLE, значение которых одинаково во всех записях."""
    if not feats:
        return {}
    first = feats[0]["properties"]
    out = {}
    for key in HOISTABLE:
        if key not in first:
            continue
        value = first[key]
        if all(f["properties"].get(key, _MISSING) == value for f in feats):
            out[key] = value
    return out


_MISSING = object()


def write_records_json(records: Iterable[ContextRecord], path: Path, *, title: str = "") -> int:
    """Пишет записи без геометрии одним JSON-массивом.

    Нужен для событий, у которых нет точки: голод по губерниям, ревизия,
    воинская повинность. В GeoJSON их не положить, а карте они нужны —
    лентой времени рядом с фактом.
    """
    rows = []
    for rec in records:
        # Пустые поля опускаются, как и в GeoJSON: у события без точки их
        # заведомо много, и файл читается вдвое легче.
        row = {k: v for k, v in rec.to_row().items() if v is not None and v != ""}
        if rec.regions:
            row["regions"] = list(rec.regions)
        if rec.extra:
            row["extra"] = rec.extra
        rows.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"name": title or path.stem, "count": len(rows), "records": rows}
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    return len(rows)


def write_xlsx(records: Sequence[ContextRecord], path: Path, *, sheet_name: str = "Данные",
               russian_headers: bool = True) -> int:
    """Пишет XLSX с русскими заголовками, как в текущих выгрузках проекта."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    # Через ту же очистку, что и многолистовая выгрузка: openpyxl запрещает
    # в имени листа `[]:*?/\`, и на двоеточии сбор падал уже после того, как
    # записал GeoJSON, — слой собран, а команда вернула ошибку. Названия
    # слоёв пишут люди, и двоеточие в них дело обычное.
    ws.title = _safe_sheet(sheet_name, wb) if sheet_name else "Данные"

    headers = [COLUMNS_RU[c] if russian_headers else c for c in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for rec in records:
        row = rec.to_row()
        ws.append([_cell(row.get(c)) for c in COLUMNS])

    widths = {"title": 42, "summary": 70, "quote": 70, "url": 45, "place_text": 34,
              "period_raw": 26, "work": 34, "actor": 26, "layer_title": 26, "uid": 24}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    wb.save(path)
    return ws.max_row - 1


def write_xlsx_multi(sheets: dict[str, Sequence[ContextRecord]], path: Path,
                     russian_headers: bool = True) -> dict[str, int]:
    """Пишет несколько слоёв в один файл — по вкладке на слой."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    counts: dict[str, int] = {}
    headers = [COLUMNS_RU[c] if russian_headers else c for c in COLUMNS]

    for name, records in sheets.items():
        ws = wb.create_sheet(title=_safe_sheet(name, wb))
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for rec in records:
            row = rec.to_row()
            ws.append([_cell(row.get(c)) for c in COLUMNS])
        widths = {"title": 42, "summary": 70, "quote": 70, "url": 45, "place_text": 34,
                  "period_raw": 26, "work": 34, "actor": 26, "layer_title": 26, "uid": 24}
        for i, col in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
        counts[name] = ws.max_row - 1

    wb.save(path)
    return counts


def write_jsonl(records: Iterable[ContextRecord], path: Path) -> int:
    """Построчный JSON — удобен для загрузки в базу и для диффов в git."""
    path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with path.open("w", encoding="utf-8") as fh:
        for rec in records:
            row = rec.to_row()
            if rec.extra:
                row["extra"] = rec.extra
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")
            n += 1
    return n


def _cell(value):
    if isinstance(value, bool):
        return "да" if value else "нет"
    return value


def _safe_sheet(name: str, wb) -> str:
    base = "".join(ch for ch in str(name) if ch not in "[]:*?/\\")[:31] or "Слой"
    title, i = base, 2
    while title in wb.sheetnames:
        suffix = f"_{i}"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    return title


# --- чтение выгрузки обратно ----------------------------------------------
# Обратная сторона write_*. Отчёт о качестве и подбор контекста работают с
# тем, что уже лежит в data/out: пересобирать три десятка слоёв из сети ради
# отчёта нельзя — половина источников за раз не отвечает, а отчёт нужен на
# каждую пересборку.


def record_from_row(row: dict) -> ContextRecord:
    """Строка выгрузки обратно в запись: JSONL, GeoJSON или `write_records_json`.

    Поле `extra` сохраняется: в нём лежат величины, ради которых слой и
    собирался, — индекс засушливости, отклонения по станции, численность.
    """
    row = dict(row)
    row["date_approx"] = row.get("date_approx") in (True, "да")
    # В таблице и в GeoJSON перечень губерний — строка «А; Б», в записи — список.
    regions = row.get("regions")
    if isinstance(regions, str):
        row["regions"] = [part.strip() for part in regions.split(";") if part.strip()]
    elif regions is None:
        row["regions"] = []
    allowed = set(ContextRecord.__dataclass_fields__)
    return ContextRecord(**{k: v for k, v in row.items() if k in allowed})


def read_jsonl(path: Path) -> list[ContextRecord]:
    """Построчный JSON обратно в записи."""
    return [record_from_row(json.loads(line))
            for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def read_records_json(path: Path) -> list[ContextRecord]:
    """Файл, записанный `write_records_json`. Чужой JSON молча пропускается.

    Рядом в каталоге лежит указатель написаний `name_variants.json` — он не
    слой, и записью притвориться не должен.
    """
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("records"), list):
        return []
    return [record_from_row(row) for row in payload["records"]
            if isinstance(row, dict) and "layer" in row]


def read_geojson(path: Path) -> list[ContextRecord]:
    """Слой карты обратно в записи.

    `write_geojson(hoist_shared=True)` выносит одинаковые у всех записей
    свойства на уровень коллекции — в поле `layer`. Без обратной склейки такой
    слой не читается вовсе: у его точек нет даже имени слоя.

    Полигон границ или чужой GeoJSON, оказавшийся в каталоге слоёв,
    пропускается: уронить сборку он не должен, а притвориться записью не может.
    """
    payload = json.loads(path.read_text(encoding="utf-8"))
    shared = payload.get("layer") or {}
    out: list[ContextRecord] = []
    for feat in payload.get("features", []):
        geometry = feat.get("geometry") or {}
        props = {**shared, **dict(feat.get("properties", {}))}
        if geometry.get("type") != "Point" or "layer" not in props:
            continue
        lon, lat = geometry["coordinates"][:2]
        props["lat"], props["lon"] = lat, lon
        out.append(record_from_row(props))
    return out


def read_layers(out_dir: Path, *, dedupe: bool = True) -> list[ContextRecord]:
    """Все собранные слои из out_dir.

    Форматов три, и одним проходом их не собрать. Точечные слои лежат в
    `geojson/`; события без точки — губернские итоги, цены, ревизии — в
    GeoJSON не попадают по определению и лежат в `*.json` рядом; полный слой
    вместе с записями без координат — в `jsonl/`. Один и тот же слой обычно
    лежит сразу в двух видах, поэтому повторы отсеиваются по `uid`.

    `dedupe=False` отдаёт всё как прочитано. Так читает отчёт о качестве: ему
    нужно видеть и совпадения `uid` у разных записей — молча схлопнутая пара
    выглядит в отчёте как одна запись, и о потере никто не узнает.
    """
    records: list[ContextRecord] = []
    seen: set[str] = set()

    def take(batch: Iterable[ContextRecord]) -> None:
        for rec in batch:
            if dedupe:
                if rec.uid in seen:
                    continue
                seen.add(rec.uid)
            records.append(rec)

    for path in sorted((out_dir / "geojson").glob("*.geojson")):
        take(read_geojson(path))
    for path in sorted(out_dir.glob("*.json")):
        take(read_records_json(path))
    for path in sorted((out_dir / "jsonl").glob("*.jsonl")):
        take(read_jsonl(path))
    return records


def read_context(out_dir: Path) -> list[ContextRecord]:
    """`context.jsonl` целиком, а при его отсутствии — собранные слои.

    Общий файл пишется последним шагом сборки и содержит всё, поэтому его
    достаточно; но он не обязателен и в репозиторий не идёт.
    """
    jsonl = out_dir / "context.jsonl"
    if jsonl.exists():
        return read_jsonl(jsonl)
    return read_layers(out_dir)
