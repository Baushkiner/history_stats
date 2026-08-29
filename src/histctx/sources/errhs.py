"""Урожай, посев и цены по губерниям: RISTAT / ERRHS.

Зачем. Погода объясняет, почему хлеб не уродился, метрическая книга —
что человек умер или ушёл из села. Между ними стоит величина, которой в
проекте до сих пор не было: сколько собрали и почём это стоило. Засуха →
недород → цена хлеба → уход на заработки и всплеск смертности. Без среднего
звена цепочка рвётся посередине, и запись «умер в 1892 году» остаётся
необъяснённой.

Откуда. `ristat.org` — Электронный архив российской исторической статистики
XVIII–XXI веков (Г. Кесслер, А. Маркевич; Международный институт социальной
истории, Амстердам). Одиннадцать наборов ERRHS по семи темам и пяти опорным
срезам: 1795, 1858, 1897, 1959, 2002 годы. Отсюда берутся четыре темы —
`4.01` (продукция сельского хозяйства в рублях), `4.02` (она же в натуре),
`7.03` (цены на землю) и `7.04` (арендная плата).

**Про кодировку значений — главное, что стоит знать.** Наборы ERRHS лежат
и в репозитории IISH (`datasets.iisg.amsterdam`), но там они закрыты: файлы
помечены `restricted`, и `/api/access/datafile/<id>` отвечает 403. В карточке
набора прямо сказано, где открыто: «Accessible at http://ristat.org». Это
удача, а не обход: выгрузка через файловый каталог `ristat.org` отдаёт
таблицу, к которой **уже применён словарь** `Vocabularies RISTAT` — территория,
показатель и единица измерения приходят словами, а не кодами. Русская версия
каталога (`/ru/catalog`) отдаёт их по-русски, и «Тамбовская губерния» из
таблицы совпадает с «Тамбовской губернией» из метрической книги без
переходного справочника. Поэтому сбор идёт с портала, а не из репозитория.

Что получается на выходе. Слой территориальный: у величины «собрано ржи в
Тамбовской губернии» координат нет и быть не может — распределять урожай
губернии по точке губернского города значило бы выдумать данные. Записи идут
со `scope = "region"`, в GeoJSON не попадают и подбираются по совпадению
губернии (`histctx.geo.region_key`).

Самая содержательная часть слоя — пары «посеяно / снято». Срезы 1795 и 1858
годов дают обе величины в одних единицах (четвертях), а их отношение — это
**сам-N**, та самая мера урожая, которой мерили современники: сам-3 значит,
что с четверти посева сняли три. После вычета семян на следующий год от
сам-3 остаётся вдвое меньше собранного, и это граница, за которой начинается
недород. Отношение считается здесь и попадает в заголовок записи: голое
«собрано 363 308 четвертей» не объясняет ничего, пока не сказано, много это
или мало.
"""

from __future__ import annotations

import html
import io
import re
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional, Sequence

from ..geo import extract_region, region_key
from ..net import RETRY_CODES, USER_AGENT, wait_for_pause
from ..schema import SCOPE_REGION, ContextRecord, LayerSpec, clean_text

HARVEST_PRICES = LayerSpec(
    slug="harvest_prices",
    title="Урожайность и хлебные цены по губерниям",
    group="economy",
    # Источник и права повторяются в каждой из десяти тысяч записей, поэтому
    # написаны коротко. Разбор расхождения между условием выгрузки и карточкой
    # набора — в docs/CATALOG.md и в журнале дискавери, а не в каждой строке.
    source="RISTAT / ERRHS, ristat.org (Г. Кесслер, А. Маркевич)",
    license=(
        "CC BY-NC-SA 4.0 (принимается при выгрузке на ristat.org); "
        "ссылка на Кесслера и Маркевича обязательна, коммерческое использование запрещено"
    ),
    description=(
        "Сколько посеяли и сколько сняли по губернии и году, поголовье скота и его падёж, "
        "цена и аренда десятины земли. Недостающее звено между погодой и записью о смерти: "
        "засуха объясняет недород, недород — цену хлеба, цена — уход на заработки."
    ),
    url="https://ristat.org/catalog",
    status="harvested",
    expected_rows=10171,
)

# Русская версия файлового каталога. Английская отдаёт то же самое, но
# названиями «Tambov governorate»: их пришлось бы сопоставлять с русскими
# вручную, а подбор контекста работает по русскому названию.
CATALOG_URL = "https://ristat.org/ru/catalog"
FORM_ID = "querytool_catalog_form"


# Темы каталога и опорные срезы, для которых они существуют. Сочетания
# сверены с чекбоксами формы: не у всякой темы есть все пять срезов.
# Срез 2002 года не берётся — он за пределами окна проекта (1800–1960),
# а объём слоя удваивает.
TOPICS: dict[str, tuple[str, tuple[int, ...]]] = {
    "4.01": ("Продукция сельского хозяйства в рублях", (1795, 1858, 1897, 1959)),
    "4.02": ("Продукция сельского хозяйства в натуре", (1795, 1858, 1897, 1959)),
    "7.03": ("Цены на землю", (1897,)),
    "7.04": ("Арендная плата за землю", (1897,)),
}

# Колонки, на которых держится разбор. Если RISTAT переименует хоть одну,
# сбор должен остановиться: пустой слой на десять тысяч записей потом
# никто не заметит.
REQUIRED_COLUMNS = ("TERRITORY", "YEAR", "VALUE", "VALUE_UNIT", "HISTCLASS1", "BASE_YEAR")

# Пропуск в таблицах ERRHS помечен точкой, а не пустой ячейкой.
MISSING = {".", "", "..", "-"}

# Ниже этого отношения «снято к посеянному» год считается недородным:
# после вычета семян на следующий сев от сам-3 остаётся вдвое меньше
# собранного. Порог грубый и служит подписью, а не выводом — само
# отношение лежит в записи числом.
POOR_YIELD = 3.0

# Единицы измерения приходят из таблицы в именительном падеже множественного
# числа («четверти», «пуды»), а в подписи их надо согласовать с числом:
# «1 пуд», «3 пуда», «38 704 пуда», «12 пудов». Три формы на единицу — один,
# два-четыре, пять и больше. Набор единиц во всех четырёх темах закрытый и
# небольшой, поэтому проще перечислить его, чем разбирать по правилам.
# Незнакомая единица пишется как в источнике: это заметно глазом и чинится
# одной строкой.
UNIT_FORMS = {
    "четверти": ("четверть", "четверти", "четвертей"),
    "пуды": ("пуд", "пуда", "пудов"),
    "головы": ("голова", "головы", "голов"),
    "рубли": ("рубль", "рубля", "рублей"),
    "марки": ("марка", "марки", "марок"),
    "десятины": ("десятина", "десятины", "десятин"),
    "кв. десятины": ("кв. десятина", "кв. десятины", "кв. десятин"),
    "тыс. десятин": ("тыс. десятин", "тыс. десятин", "тыс. десятин"),
    "гектары": ("гектар", "гектара", "гектаров"),
    "гектолитры": ("гектолитр", "гектолитра", "гектолитров"),
    "килограммы": ("килограмм", "килограмма", "килограммов"),
    "тонны": ("тонна", "тонны", "тонн"),
    "центнеры": ("центнер", "центнера", "центнеров"),
    "штуки": ("штука", "штуки", "штук"),
    "человек": ("человек", "человека", "человек"),
    "копны": ("копна", "копны", "копён"),
    "возы": ("воз", "воза", "возов"),
    "ведро": ("ведро", "ведра", "вёдер"),
    "берковцы": ("берковец", "берковца", "берковцев"),
    "кв. версты": ("кв. верста", "кв. версты", "кв. вёрст"),
    "кв. километры": ("кв. километр", "кв. километра", "кв. километров"),
    "кубические сажени": ("кубическая сажень", "кубические сажени",
                          "кубических саженей"),
    # «число» — не единица, а пометка «просто количество»: в подписи она лишняя.
    "число": ("", "", ""),
}

# Пометка записи, чью территорию подбор по названию может не найти.
# Такие записи остаются в слое — выбрасывать их молча нельзя, см. is_matchable.
CONFIDENCE_UNMATCHED = "region_unmatched"

_RE_BUILD_ID = re.compile(r'name="form_build_id"\s+value="([^"]+)"')
_RE_SUBMIT = re.compile(r'<input[^>]*id="edit-submit"[^>]*value="([^"]*)"')
_RE_ZIP_URL = re.compile(
    r'(https://etl\.ristat\.org/service/filecatalogget\?zip=[^"\'\s<>]+)')

# Уточнение в скобках у названия территории: «Астраханская губерния (1802>)».
# Для подбора оно лишнее, но из исходного текста не выбрасывается.
_RE_QUALIFIER = re.compile(r"\s*\([^)]*\)\s*$")

# «рожь озимая в посеве», «хлеб яровой (снято)» — показатель и роль в одной
# строке. Латинская «o» в «(посеянo)» не опечатка разбора, а опечатка
# источника: в срезе 1858 года она встречается у капусты.
_RE_ROLE = re.compile(
    r"\s*(?:в\s+(посеве|урожае)|\((посеян[оo]|снят[оo])\))\s*$",
    re.IGNORECASE,
)
_ROLE_BY_WORD = {"посеве": "посев", "урожае": "урожай",
                 "посеяно": "посев", "снято": "урожай"}


class ErrhsError(RuntimeError):
    """Выгрузка RISTAT не получена или пришла не в том виде, какого мы ждём."""


@dataclass(frozen=True)
class Figure:
    """Одна величина таблицы ERRHS: что, где, когда, сколько и в чём."""

    territory: str          # как написано в таблице, с уточнением в скобках
    region: str             # то же без уточнения — по нему идёт подбор
    year: Optional[int]
    value: float
    unit: Optional[str]
    subject: str            # показатель без «в посеве» / «(снято)»
    role: Optional[str]     # «посев», «урожай» или None
    topic: str              # тема каталога: 4.01, 4.02, 7.03, 7.04
    benchmark: int          # опорный срез, из которого взята таблица
    row_id: str
    district: Optional[str] = None
    source_book: Optional[str] = None
    page: Optional[str] = None
    comment: Optional[str] = None
    year_is_benchmark: bool = False   # год не указан, взят из опорного среза


@dataclass
class RistatCatalog:
    """Клиент файлового каталога ristat.org.

    Выгрузка идёт в три шага, и обойти их нельзя: каталог отдаёт не файл по
    постоянному адресу, а архив, собранный под запрос. Сначала со страницы
    берётся `form_build_id`, потом форма отправляется с отметками нужных тем,
    и только в ответе появляется ссылка на архив. Ссылка одноразовая, поэтому
    скачанное кладётся в кэш: пересобирать слой, дёргая сайт заново, незачем.
    """

    timeout: int = 180
    max_retries: int = 4
    pause_sec: float = 1.0
    cache_dir: Optional[Path] = None
    _last_call: float = field(default=0.0, repr=False)

    def table(self, topic: str, benchmark: int) -> tuple[list[str], list[dict]]:
        """Заголовок и строки таблицы одной темы за один опорный срез."""
        return read_table(self.zip_bytes(topic, benchmark), what=f"{topic}/{benchmark}")

    def zip_bytes(self, topic: str, benchmark: int) -> bytes:
        cached = self._cached(topic, benchmark)
        if cached is not None and cached.exists():
            return cached.read_bytes()

        page = self._open(CATALOG_URL).decode("utf-8", "replace")
        build_id = _first(_RE_BUILD_ID, page)
        if not build_id:
            raise ErrhsError(
                f"на странице {CATALOG_URL} нет поля form_build_id — форма каталога "
                "изменилась, разбор надо править"
            )
        # Значение кнопки берётся со страницы: оно зависит от языка версии.
        submit = _first(_RE_SUBMIT, page) or "download"
        payload = urllib.parse.urlencode({
            "form_build_id": build_id,
            "form_id": FORM_ID,
            "op": submit,
            f"{topic.replace('.', '_')}_{benchmark}": "1",
        }).encode("utf-8")

        # Ответ разэкранируется: ссылка на архив приходит внутри HTML, и
        # «&amp;» в ней превратился бы в параметр с именем «amp;…».
        answer = html.unescape(self._open(CATALOG_URL, data=payload).decode("utf-8", "replace"))
        url = _first(_RE_ZIP_URL, answer)
        if not url:
            raise ErrhsError(
                f"каталог не дал ссылки на архив для темы {topic} за {benchmark} год. "
                "Обычная причина — такого сочетания темы и среза не существует; "
                f"известные срезы темы: {TOPICS.get(topic, ('', ()))[1]}"
            )
        raw = self._open(url)
        if not raw.startswith(b"PK"):
            raise ErrhsError(
                f"по ссылке {url} пришёл не zip, а {len(raw)} байт, начинающихся с "
                f"{raw[:16]!r}"
            )
        if cached is not None:
            cached.parent.mkdir(parents=True, exist_ok=True)
            cached.write_bytes(raw)
        return raw

    def _cached(self, topic: str, benchmark: int) -> Optional[Path]:
        if self.cache_dir is None:
            return None
        return Path(self.cache_dir) / f"errhs_{topic.replace('.', '_')}_{benchmark}.zip"

    def _open(self, url: str, data: Optional[bytes] = None) -> bytes:
        req = urllib.request.Request(url, data=data, headers={
            "User-Agent": USER_AGENT,
            "Accept": "*/*",
        })
        delay = 2.0
        last: Optional[Exception] = None
        for attempt in range(self.max_retries):
            self._throttle()
            try:
                with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                    return resp.read()
            except urllib.error.HTTPError as exc:
                last = exc
                if exc.code not in RETRY_CODES:
                    body = exc.read()[:300].decode("utf-8", "replace")
                    raise ErrhsError(f"HTTP {exc.code} на {url}: {body}") from exc
            except (urllib.error.URLError, TimeoutError, ConnectionError) as exc:
                last = exc
            if attempt < self.max_retries - 1:
                time.sleep(delay)
                delay *= 2
        raise ErrhsError(f"{url}: не удалось получить ответ за {self.max_retries} попыток: {last}")

    def _throttle(self) -> None:
        """Пауза между запросами: каталог собирает архив на лету, и торопить его нечем."""
        self._last_call = wait_for_pause(self._last_call, self.pause_sec)


def read_table(raw: bytes, *, what: str = "выгрузка") -> tuple[list[str], list[dict]]:
    """Достаёт таблицу из архива каталога.

    В архиве лежат сама таблица (XLSX) и документация к ней (PDF с описанием
    среза, классификации и границ территорий). Документация в слой не идёт,
    но скачивается вместе с данными — это условие выгрузки, а не наше решение.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            names = [n for n in archive.namelist() if n.lower().endswith(".xlsx")]
            if not names:
                raise ErrhsError(
                    f"{what}: в архиве нет XLSX, только {sorted(archive.namelist())}"
                )
            data = archive.read(names[0])
    except zipfile.BadZipFile as exc:
        raise ErrhsError(f"{what}: архив не читается ({exc})") from exc

    header, rows = read_sheet(data, what=f"{what}: {names[0]}")
    return header, rows


def read_sheet(data: bytes, *, what: str = "таблица") -> tuple[list[str], list[dict]]:
    """Читает первый лист XLSX в список словарей."""
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    stream = sheet.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(stream)]
    except StopIteration:
        raise ErrhsError(f"{what}: пустой лист") from None
    check_columns(header, what=what)
    rows = [dict(zip(header, row)) for row in stream]
    book.close()
    return header, rows


def check_columns(header: Sequence[str], *, what: str = "таблица") -> None:
    """Останавливает сбор, если в таблице нет колонок, на которых держится разбор."""
    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise ErrhsError(
            f"{what}: нет обязательных колонок {missing}; в таблице: {list(header)}. "
            "Формат выгрузки RISTAT изменился — поправьте разбор, прежде чем собирать."
        )


def read_figures(rows: Iterable[dict], *, topic: str, benchmark: int) -> list[Figure]:
    """Приводит строки таблицы к величинам.

    Строка без территории или без числа в слой не идёт: показать на карте
    «Тамбовская губерния, 1794, значение неизвестно» нельзя, а сколько таких
    строк было — печатает сборщик.
    """
    out: list[Figure] = []
    for row in rows:
        territory = _cell(row, "TERRITORY")
        value = _number(row.get("VALUE"))
        if not territory or value is None:
            continue
        subject, role = subject_and_role(row)
        if not subject:
            continue

        year = _int(row.get("YEAR"))
        from_benchmark = year is None
        out.append(Figure(
            territory=territory,
            region=normalize_region(territory),
            year=year if year is not None else benchmark,
            value=value,
            unit=_cell(row, "VALUE_UNIT"),
            subject=subject,
            role=role,
            topic=topic,
            benchmark=benchmark,
            # Запасной номер помечен словом: столкнуться с настоящим ID
            # он не может, а uid записи держится именно на нём.
            row_id=_cell(row, "ID") or f"row{len(out) + 1}",
            district=_cell(row, "DISTRICT"),
            source_book=_cell(row, "SOURCE"),
            page=_cell(row, "PAGE"),
            comment=_cell(row, "COMMENT_NABORSHIK") or _cell(row, "COMMENT_SOURCE"),
            year_is_benchmark=from_benchmark,
        ))
    return out


def region_records(figures: Iterable[Figure], *, spec: LayerSpec = HARVEST_PRICES,
                   url: Optional[str] = None) -> list[ContextRecord]:
    """Территориальные записи слоя: губерния, год, величина.

    Посев и урожай одной культуры в одной губернии за один год сводятся в
    одну запись с отношением сам-N. Пара считается только когда в срезе ровно
    одна строка посева и ровно одна строка урожая: у нескольких строк на один
    показатель источники разные, и складывать их — уже реконструкция, а не
    выгрузка. Несведённые строки идут в слой по отдельности.
    """
    paired, singles = _pair_sowing_and_harvest(figures)
    out: list[ContextRecord] = []
    for harvest, sowing in paired:
        out.append(_record(harvest, spec=spec, url=url, sowing=sowing))
    for figure in singles:
        out.append(_record(figure, spec=spec, url=url))
    out.sort(key=lambda r: (r.region or "", r.year_from or 0, r.title))
    return out


def category_of(figure: Figure) -> str:
    """Что это за величина — для легенды и фильтров карты."""
    if figure.topic == "7.03":
        return "цена земли"
    if figure.topic == "7.04":
        return "аренда земли"
    if figure.topic == "4.01":
        return "выпуск в рублях"
    if "падеж" in figure.subject.lower() or "падёж" in figure.subject.lower():
        return "падёж скота"
    unit = (figure.unit or "").lower()
    if unit in {"головы", "штуки", "число"}:
        return "поголовье"
    if unit in {"десятины", "гектары", "тыс. десятин", "кв. версты", "кв. километры",
                "кв. десятины"}:
        return "угодья"
    if figure.role == "посев":
        return "посев"
    return "урожай"


def is_matchable(region: str) -> bool:
    """Найдёт ли подбор контекста эту территорию по названию.

    Подбор сравнивает ключи `histctx.geo.region_key` — у факта и у записи.
    Ключ факта берётся из текста документа: он либо задан прямо, либо вынут
    оттуда `extract_region`. Название проходит в двух случаях:

    * его узнаёт извлекатель — «Тамбовская губерния», «Акмолинская область»;
    * его ключ состоит из одной основы — «Татарская АССР» даёт «татарск», и
      документ, как бы он ни назвал эту территорию, даст тот же ключ.

    Не проходят составные названия: «Земля войска Донского» даёт ключ
    «земля донск», а «Адыгейская автономная область Краснодарского края» —
    «адыгейск автономн краснодарск». Воспроизвести такой ключ из документа
    нечем, и запись помечается `region_unmatched`. Из слоя она не убирается:
    величина верна, ненадёжна только привязка по названию.
    """
    key = region_key(region)
    return bool(key) and (extract_region(region) is not None or " " not in key)


def normalize_region(territory: str) -> str:
    """«Астраханская губерния (1802>)» → «Астраханская губерния».

    Уточнение в скобках говорит, о какой редакции губернии речь, и для
    истории важно, но подбору мешает: одна и та же губерния приходит в срезе
    двумя написаниями. Исходный текст сохраняется в `place_text`.
    """
    return _RE_QUALIFIER.sub("", territory).strip() or territory


def subject_and_role(row: dict) -> tuple[str, Optional[str]]:
    """Собирает показатель из цепочки HISTCLASS и отделяет от него роль.

    Уточнения в цепочке — не украшение: в срезе 1858 года «хлеб озимой
    (посеяно)» встречается трижды в одной губернии, и различает эти строки
    только вторая ступень — у помещиков, у государственных крестьян, в уделе.
    Слить их в один показатель значило бы сложить несопоставимое и потерять
    все три пары «посеяно / снято».
    """
    head, role = split_role(_cell(row, "HISTCLASS1") or "")
    parts = [head] if head else []
    for level in range(2, 11):
        value = _cell(row, f"HISTCLASS{level}")
        if value:
            parts.append(value)
    return ", ".join(parts), role


def split_role(indicator: str) -> tuple[str, Optional[str]]:
    """«рожь озимая в посеве» → («рожь озимая», «посев»)."""
    match = _RE_ROLE.search(indicator)
    if not match:
        return indicator.strip(), None
    word = (match.group(1) or match.group(2) or "").lower().replace("o", "о")
    return indicator[: match.start()].strip(), _ROLE_BY_WORD.get(word)


def yield_ratio(harvest: Figure, sowing: Figure) -> Optional[float]:
    """Сам-N: во сколько раз снятое больше посеянного."""
    if sowing.value <= 0 or harvest.unit != sowing.unit:
        return None
    return harvest.value / sowing.value


def _pair_sowing_and_harvest(
    figures: Iterable[Figure],
) -> tuple[list[tuple[Figure, Figure]], list[Figure]]:
    groups: dict[tuple, list[Figure]] = {}
    loose: list[Figure] = []
    for figure in figures:
        if figure.role is None:
            loose.append(figure)
            continue
        # Тема и опорный срез входят в ключ: пара «посеяно / снято» имеет
        # смысл только внутри одной таблицы. Свести строку из среза 1795 года
        # со строкой из среза 1858-го значило бы посчитать отношение по двум
        # разным источникам и приписать его одному из них.
        groups.setdefault(
            (figure.topic, figure.benchmark, figure.region, figure.year,
             figure.subject, figure.unit), []
        ).append(figure)

    paired: list[tuple[Figure, Figure]] = []
    for group in groups.values():
        sown = [f for f in group if f.role == "посев"]
        reaped = [f for f in group if f.role == "урожай"]
        # Сводить в одну запись имеет смысл только тогда, когда получается
        # сам-N. Если посеяно ноль (в источнике культуры ещё нет) или единицы
        # разные, обе строки идут в слой порознь: слить их — значит потерять
        # одну из двух, а это то самое молчаливое удаление, которого нельзя.
        # Проверка именно на None: полный неурожай даёт сам-0, и это
        # не «пары нет», а самая говорящая запись во всём слое.
        if len(sown) == 1 and len(reaped) == 1 and (
                yield_ratio(reaped[0], sown[0]) is not None):
            paired.append((reaped[0], sown[0]))
        else:
            loose.extend(group)
    return paired, loose


def _record(figure: Figure, *, spec: LayerSpec, url: Optional[str],
            sowing: Optional[Figure] = None) -> ContextRecord:
    ratio = yield_ratio(figure, sowing) if sowing is not None else None
    subject = figure.subject
    measure = _measure(figure.value, figure.unit)

    if ratio is not None:
        head = f"урожай сам-{_ratio_text(ratio)}"
    elif figure.role == "посев":
        head = f"посеяно {measure}"
    elif figure.role == "урожай":
        head = f"снято {measure}"
    else:
        head = measure

    title = f"{_capitalize(subject)}: {head} — {figure.region}, {figure.year}"

    # Название темы в extra не пишется — только код: расшифровка одна на слой
    # и лежит в TOPICS, а в файле она повторилась бы десять тысяч раз.
    # Первоисточник тоже не дублируется: он назван в описании записи.
    extra = {
        "тема": figure.topic,
        "опорный_срез": figure.benchmark,
        "показатель": subject,
        "значение": figure.value,
        "единица": figure.unit,
        "страница": figure.page,
        "примечание": figure.comment,
    }
    if sowing is not None:
        extra["посеяно"] = sowing.value
        if ratio is not None:
            extra["сам"] = round(ratio, 2)

    matched = is_matchable(figure.region)
    return spec.new_record(
        title=title,
        category=category_of(figure),
        scope=SCOPE_REGION,
        # Исходный текст места сохраняется только тогда, когда он отличается
        # от губернии: у большинства строк это одна и та же строка.
        place_text=figure.territory if figure.territory != figure.region else None,
        region=figure.region,
        regions=[figure.region],
        district=figure.district,
        year_from=figure.year,
        year_to=figure.year,
        date_precision="year",
        date_approx=figure.year_is_benchmark,
        period_raw=(f"опорный срез {figure.benchmark} года"
                    if figure.year_is_benchmark else str(figure.year)),
        summary=_summary(figure, sowing, ratio),
        url=url or spec.url,
        source_id=f"{figure.topic}:{figure.benchmark}:{figure.row_id}",
        confidence="ok" if matched else CONFIDENCE_UNMATCHED,
        extra={k: v for k, v in extra.items() if v not in (None, "")},
    )


def unit_text(unit: Optional[str], value: float) -> str:
    """Единица, согласованная с числом: «1 пуд», «3 пуда», «12 пудов».

    У дробного числа форма всегда «два-четыре»: «1,9 рубля», а не «1,9 рублей».
    """
    if not unit:
        return ""
    forms = UNIT_FORMS.get(unit.strip().lower())
    if forms is None:
        return unit
    if abs(value - round(value)) > 1e-9:
        return forms[1]
    whole = abs(int(round(value)))
    if 11 <= whole % 100 <= 14:
        return forms[2]
    last = whole % 10
    if last == 1:
        return forms[0]
    return forms[1] if 2 <= last <= 4 else forms[2]


def _summary(figure: Figure, sowing: Optional[Figure], ratio: Optional[float]) -> str:
    parts = [f"{figure.region}, {figure.year} год: {figure.subject} — "
             f"{_measure(figure.value, figure.unit)}"]
    if sowing is not None:
        parts.append(f"посеяно {_measure(sowing.value, sowing.unit)}")
    if ratio is not None:
        parts.append(f"урожай сам-{_ratio_text(ratio)}")
        if ratio < POOR_YIELD:
            # Ниже сам-3 после вычета семян на следующий сев остаётся вдвое
            # меньше собранного — это и есть недород, который через год
            # виден в третьей части метрической книги.
            parts.append("недород: после вычета семян на новый сев остаётся "
                         "вдвое меньше собранного")
    if figure.year_is_benchmark:
        parts.append(f"год в источнике не указан, взят опорный срез {figure.benchmark}")
    # Примечание составителя в описание не идёт: оно бывает в полстраницы и
    # вытесняет собой саму величину. Оно лежит целиком в extra.
    tail = f" Источник: {figure.source_book.rstrip('.')}." if figure.source_book else ""
    return "; ".join(parts) + "." + tail


def _measure(value: float, unit: Optional[str]) -> str:
    """«363 308 четвертей», «1,9 рубля», «18» — число с согласованной единицей."""
    return f"{_amount(value)} {unit_text(unit, value)}".strip()


def _amount(value: float) -> str:
    """Число с разрядами: «363 308», «1,9»."""
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}".replace(",", " ")
    return f"{value:,.2f}".replace(",", " ").replace(".", ",").rstrip("0").rstrip(",")


def _ratio_text(ratio: float) -> str:
    return f"{ratio:.1f}".replace(".", ",")


def _capitalize(text: str) -> str:
    return text[:1].upper() + text[1:] if text else text


def _first(pattern: re.Pattern, text: str) -> Optional[str]:
    match = pattern.search(text)
    return match.group(1) if match else None


def _cell(row: dict, key: str) -> Optional[str]:
    """Значение ячейки или None; пропуск в ERRHS помечен точкой."""
    text = clean_text(row.get(key))
    return None if text is None or text in MISSING else text


def _int(value) -> Optional[int]:
    """Свой, а не `normalize.to_int`: опорный срез в таблицах ERRHS записан
    дробным («1897.0»), и строгий разбор его не признаёт."""
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _number(value) -> Optional[float]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip() if value is not None else ""
    if not text or text in MISSING:
        return None
    try:
        return float(text.replace(" ", "").replace(" ", "").replace(",", "."))
    except ValueError:
        return None
