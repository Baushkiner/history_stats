"""Оформление книг XLSX, которые читают глазами.

Таких выгрузок две — общая (`scripts/export_layer_xlsx.py`) и своя у слоя
лагерей (`scripts/export_gulag_xlsx.py`), — и рисуют они одно и то же: шапку
белым по тёмному, заголовки разделов, шапки колонок внутри листа, пары
«ключ — значение» на листе об источнике. Оформление лежит здесь, чтобы книги
выглядели одинаково: пока каждая выгрузка держала свою копию, правка в одной
до другой не доходила.

Записью слоя как таблицы ведает `io_formats.write_xlsx`: там плоский вывод
для машины, здесь — оформление для человека.

openpyxl подтягивается внутри функций, а не на уровне модуля: пакет нужен
только выгрузкам, а `histctx` читается и без него.
"""

from __future__ import annotations

FONT = "Arial"
SIZE = 10

HEAD_BG = "44546A"          # тёмно-синий фон шапки
HEAD_FG = "FFFFFF"
LINK_FG = "0563C1"          # ссылка — как в самом Excel
WARN_FG = "C00000"          # запись, о которой читателя предупреждают
NOTE_FG = "595959"          # пояснение под заголовком раздела


def new_workbook():
    """Пустая книга проекта: без листа по умолчанию, со шрифтом и пересчётом."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    # openpyxl кладёт формулы строками без значений: без этой пометки сводка
    # откроется пустой, пока читатель не нажмёт пересчёт руками.
    wb.calculation.fullCalcOnLoad = True
    # Шрифт задаётся книге целиком, а не каждой ячейке отдельно: на слое
    # храмов (31 896 записей) поячеечная простановка — это две трети миллиона
    # обращений к стилям, и выгрузка не укладывалась в две минуты.
    #
    # Менять нужно именно нулевой шрифт списка: ячейка без своего оформления
    # ссылается на `cellXfs[0]`, а тот жёстко указывает `fontId="0"`. Правки
    # одного лишь именованного стиля «Normal» ячейкам не видно — проверено по
    # `xl/styles.xml`, там оставался Calibri.
    # Выравнивание так задать нельзя — `cellXfs[0]` пишется без `applyAlignment`,
    # и ячейка остаётся с умолчанием. Строки таблицы однострочные, разницы не
    # видно; там, где она важна (шапки, сводка), выравнивание ставится явно.
    wb._fonts[0] = Font(name=FONT, size=SIZE)
    wb._named_styles["Normal"].font = Font(name=FONT, size=SIZE)
    wb.remove(wb.active)
    return wb


def head_row(ws, titles, *, height: int = 32) -> None:
    """Первая строка листа: белым по тёмному, с переносом."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append(list(titles))
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, color=HEAD_FG, size=SIZE)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = height


def cell_value(value):
    """Значение так, как его читает человек: «да»/«нет» вместо True/False."""
    import json

    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def link_font():
    """Шрифт ссылки — ставится поверх книжного, по ячейке."""
    from openpyxl.styles import Font
    return Font(name=FONT, size=SIZE, color=LINK_FG, underline="single")


def warn_font():
    """Шрифт пометки: запись оставлена, но читателя о ней предупреждают."""
    from openpyxl.styles import Font
    return Font(name=FONT, size=SIZE, color=WARN_FG)


class Cursor:
    """Пишет лист сверху вниз и помнит, на какой строке стоит.

    Номер строки нужен самим формулам («сколько записей со значением из A12»),
    а вычислять его из `ws.max_row` нельзя: пустая строка-отбивка счётчик
    openpyxl не двигает, и ссылки разъезжаются с данными.
    """

    def __init__(self, ws):
        from openpyxl.styles import Font
        self.ws = ws
        self.line = 0
        self.body = Font(name=FONT, size=SIZE)
        self.bold = Font(name=FONT, bold=True, size=SIZE)
        self.title = Font(name=FONT, bold=True, size=13)
        self.note = Font(name=FONT, italic=True, size=9, color=NOTE_FG)
        self.head = Font(name=FONT, bold=True, size=SIZE, color=HEAD_FG)

    def row(self, values, *, font=None, formats=None, height=None) -> int:
        """Строка значений: подпись слева, числа справа. Возвращает её номер."""
        from openpyxl.styles import Alignment
        self.line += 1
        for i, value in enumerate(values, start=1):
            cell = self.ws.cell(row=self.line, column=i, value=value)
            cell.font = font or self.body
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if i == 1 else "right")
            if formats and i - 1 < len(formats) and formats[i - 1]:
                cell.number_format = formats[i - 1]
        if height:
            self.ws.row_dimensions[self.line].height = height
        return self.line

    def blank(self) -> None:
        """Отбивка между разделами."""
        self.line += 1

    def section(self, text, comment=None) -> None:
        """Заголовок раздела и, если есть, пояснение под ним."""
        self.row([text], font=self.title, height=26)
        if comment:
            self.row([comment], font=self.note, height=14)

    def header(self, cells, *, height: int = 28) -> int:
        """Шапка колонок внутри листа — та же, что и у листа целиком."""
        from openpyxl.styles import Alignment, PatternFill
        line = self.row(cells, font=self.head)
        for i in range(1, len(cells) + 1):
            cell = self.ws.cell(row=line, column=i)
            cell.fill = PatternFill("solid", fgColor=HEAD_BG)
            cell.alignment = Alignment(vertical="center", horizontal="center",
                                       wrap_text=True)
        self.ws.row_dimensions[line].height = height
        return line

    def pair(self, key, value, *, per_line: int = 100) -> int:
        """Пара «ключ — значение» листа об источнике: подпись слева, текст справа.

        Высота считается по длине текста: openpyxl не умеет подгонять строку
        под перенос, а без этого длинное значение обрезается видимой частью.
        `per_line` — сколько знаков влезает в ширину колонки со значением.
        """
        from openpyxl.styles import Alignment
        line = self.row([key, str(value)], font=self.bold)
        self.ws.cell(row=line, column=1).alignment = Alignment(vertical="top")
        val = self.ws.cell(row=line, column=2)
        val.font = self.body
        val.alignment = Alignment(wrap_text=True, vertical="top")
        self.ws.row_dimensions[line].height = 14 * (1 + len(str(value)) // per_line)
        return line

    def paragraph(self, text, *, per_line: int = 130) -> int:
        """Абзац на всю ширину листа об источнике."""
        from openpyxl.styles import Alignment
        line = self.row([text], font=self.body)
        cell = self.ws.cell(row=line, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        self.ws.merge_cells(start_row=line, start_column=1, end_row=line, end_column=2)
        self.ws.row_dimensions[line].height = 14 * (1 + len(str(text)) // per_line)
        return line
