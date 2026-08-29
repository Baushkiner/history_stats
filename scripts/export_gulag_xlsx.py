#!/usr/bin/env python3
"""Выгрузка слоя лагерей ГУЛАГа в один XLSX — для чтения глазами, не машиной.

    python3 scripts/export_gulag_xlsx.py                       # data/out/xlsx/gulag_camps.xlsx
    python3 scripts/export_gulag_xlsx.py --out путь/файл.xlsx

Зачем отдельная выгрузка. Общий `write_xlsx` пишет поля единой схемы и
выбрасывает `extra`, а у этого слоя в `extra` лежит самое интересное:
учтённая численность заключённых по годам, наибольшая численность и
региональная группировка проекта. Здесь эти числа разворачиваются в
отдельный лист и сводку, а вид работ вынимается обратно из `summary` —
в записи он отдельным полем не хранится.

Четыре листа:

* «Лагеря» — по строке на место размещения управления: название, тип, вид
  работ, место, годы, наибольшая численность, координата, ссылка на карточку;
* «Численность по годам» — та же строка, развёрнутая по годам 1918–1960;
* «Сводка» — счёт по типам, регионам и видам работ и учтённая численность по
  годам; все числа считаются формулами, а не вписаны;
* «Об источнике» — откуда данные, что можно с ними делать, чего в них нет,
  и словарь колонок.

Числа не пересчитываются при записи: openpyxl кладёт формулы строками без
значений. Пересчёт — LibreOffice, см. `docs/HARVEST.md`.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SRC = ROOT / "data" / "out" / "geojson" / "gulag_camps.geojson"
DEFAULT_OUT = ROOT / "data" / "out" / "xlsx" / "gulag_camps.xlsx"

# Справочник «Мемориала» доведён до 1960-го, первые лагеря — 1918 год.
YEAR_MIN, YEAR_MAX = 1918, 1960

# «Исправительно-трудовые лагеря; вид работ — лесозаготовки; наибольшая…»
_RE_ACTIVITY = re.compile(r"^вид работ — (.+)$")

FONT = "Arial"

# Колонка → (заголовок, ширина). Порядок здесь — порядок в листе «Лагеря».
CAMP_COLUMNS = [
    ("no",            "№",                                6),
    ("title",         "Название",                        44),
    ("category",      "Тип лагеря",                      30),
    ("activity",      "Вид работ",                       30),
    ("place_text",    "Место (как в источнике)",         52),
    ("region",        "Губерния/область",                24),
    ("district",      "Уезд/район",                      20),
    ("camp_region",   "Регион (группировка проекта)",    30),
    ("period_raw",    "Годы работы",                     14),
    ("year_from",     "Год от",                           9),
    ("year_to",       "Год до",                           9),
    ("date_approx",   "Датировка приблизительна",        14),
    ("peak",          "Наибольшая учтённая численность", 16),
    ("peak_year",     "Год наибольшей численности",      13),
    ("years_known",   "Лет с числами",                   10),
    ("lat",           "Широта",                          11),
    ("lon",           "Долгота",                         11),
    ("url",           "Карточка проекта",                30),
    ("published",     "Карточка опубликована",           14),
    ("summary",       "Описание из слоя",                70),
    ("camp_id",       "ID управления",                   11),
    ("first_of_camp", "Первая запись управления",        12),
    ("uid",           "UID записи",                      26),
    ("source_id",     "ID в источнике",                  14),
]

ABOUT = [
    ("h", "Лагеря и лагерные управления ГУЛАГа"),
    ("p", "Выгрузка слоя `gulag_camps` из репозитория исторического контекста "
          "genealogy/history_stats. Файл пересобирается командой "
          "`python3 scripts/export_gulag_xlsx.py`, данные слоя — "
          "`python3 scripts/harvest_gulag.py --build`."),
    ("h", "Откуда данные"),
    ("kv", "Источник", "«Карта ГУЛАГа» Музея истории ГУЛАГа "
                       "(https://gulagmap.ru/), составлена по справочнику "
                       "«Мемориала» «Система исправительно-трудовых лагерей "
                       "в СССР, 1923–1960» — по фондам ГАРФ"),
    ("kv", "Как забрано", "Открытый JSON API проекта: gulagmap.ru/api/camps "
                          "и три справочника к нему (типы лагерей, виды работ, регионы)"),
    ("kv", "Права", "Открытых условий проект не публикует. Поэтому взяты только "
                    "факты — название, координата, годы работы, тип, вид работ, "
                    "учтённая численность — и ссылка на карточку. Авторская "
                    "историческая справка карточки и фотографии не копируются: "
                    "за ними строка отсылает на gulagmap.ru. Колонка «Описание "
                    "из слоя» — фраза, собранная из справочников и чисел, а не "
                    "пересказ справки музея."),
    ("h", "Что нужно знать, прежде чем считать"),
    ("p", "1. Строка — это место размещения управления, а не лагерь. "
          "Управление, переезжавшее с места на место, даёт по строке на каждое "
          "место: Северо-Двинский ИТЛ стоял сначала в Котласе, потом в "
          "Архангельске, и для поиска личного дела это разные адреса. Колонка "
          "«Первая запись управления» ставит 1 у первой строки каждого "
          "управления — по ней считается число самих управлений."),
    ("p", "2. Часть карточек проект у себя не показывает. API их отдаёт, и по "
          "составу полей они не хуже прочих, но выдавать их за проверенные "
          "нельзя: в колонке «Карточка опубликована» у них стоит «нет». "
          "Молча выброшены они не были — правило репозитория."),
    ("p", "3. Численность — учтённая, а не действительная. Это цифры отчётности "
          "лагерных управлений: сколько человек числилось на дату учёта. "
          "Пустая ячейка на листе «Численность по годам» значит «числа нет», "
          "ноль значит «в отчёте ноль». Сумма по году на листе «Сводка» — "
          "сумма того, что учтено в этих карточках, а не численность системы "
          "лагерей в этом году."),
    ("p", "4. Годы работы часто приблизительны. Где датировки в описании места "
          "не было, границы взяты по годам, за которые есть численность; такая "
          "строка помечена в колонке «Датировка приблизительна». Рамка "
          "1918–1960 — границы справочника, а не границы явления."),
    ("p", "5. Вид работ вынут обратно из текста описания: отдельным полем слой "
          "его не хранит. Написание вроде «хозяйственные Работы» — из "
          "справочника проекта, не опечатка выгрузки."),
    ("p", "6. Одна карточка похожа на пробную запись проекта: «Новый лагерь», "
          "1939 год, 456 789 человек — число выше любого из опубликованных "
          "и стоит в неопубликованной карточке. Молча выброшена она не была: "
          "строка на месте, а у ячейки с числом висит примечание. Проверка "
          "общая, не по названию, — если проект заведёт такую же ещё раз, "
          "выгрузка пометит и её."),
    ("h", "Словарь колонок листа «Лагеря»"),
]

SUSPICIOUS_NOTE = (
    "Число больше, чем в любой опубликованной карточке, а сама карточка "
    "проектом не опубликована. Похоже на пробную запись: проверяйте по "
    "карточке на gulagmap.ru, прежде чем считать. Из выгрузки строка не "
    "убрана — ничего не удаляется молча."
)

COLUMN_NOTES = {
    "Название": "Название управления, в скобках — сокращение: «Джугджурский ИТЛ (Джугджурлаг)»",
    "Тип лагеря": "Справочник проекта: ИТЛ, лагеря ГУПВИ, концлагеря Гражданской войны, "
                  "проверочно-фильтрационные, особые лагеря, лагпункты, спецпоселения",
    "Вид работ": "Справочник проекта: производственный профиль лагеря",
    "Место (как в источнике)": "Текст описания места из карточки, с исторической и нынешней привязкой",
    "Губерния/область": "Разобрано из текста места, где разбор удался",
    "Уезд/район": "Разобрано из текста места, где разбор удался",
    "Регион (группировка проекта)": "Шестнадцать крупных регионов, которыми проект делит карту",
    "Годы работы": "Как записано в слое: один год или период",
    "Год от": "Начало работы на этом месте",
    "Год до": "Конец работы на этом месте",
    "Датировка приблизительна": "«да» — либо период, либо годы взяты по данным о численности",
    "Наибольшая учтённая численность": "Наибольшее из чисел по годам; нули в расчёт не идут",
    "Год наибольшей численности": "Год, в котором учтено это наибольшее число",
    "Лет с числами": "Сколько лет по этому месту вообще есть цифра — мера полноты строки",
    "Широта": "Градусы, WGS 84",
    "Долгота": "Градусы, WGS 84",
    "Карточка проекта": "Ссылка на карточку gulagmap.ru — там историческая справка и источники",
    "Карточка опубликована": "«нет» — карточка есть в API, но проект её у себя не показывает",
    "Описание из слоя": "Фраза, собранная из справочников и чисел (не текст музея)",
    "ID управления": "Идентификатор управления в API проекта; у переезжавшего лагеря повторяется",
    "Первая запись управления": "1 у первой строки управления — чтобы считать управления, а не места",
    "UID записи": "Идентификатор строки в слое репозитория",
    "ID в источнике": "«управление:место» в API проекта",
}


def load(path: Path) -> list[dict]:
    """Читает слой и раскладывает свойства по колонкам выгрузки."""
    with path.open(encoding="utf-8") as fh:
        payload = json.load(fh)
    features = payload.get("features") or []
    if not features:
        raise SystemExit(f"{path}: в слое нет записей — сначала соберите его "
                         f"командой scripts/harvest_gulag.py --build")

    rows = []
    for feature in features:
        props = feature.get("properties") or {}
        extra = props.get("extra") or {}
        lon, lat = (feature.get("geometry") or {}).get("coordinates", [None, None])[:2]
        prisoners = {int(y): v for y, v in (extra.get("prisoners") or {}).items()
                     if str(y).isdigit()}
        peak, peak_year = _peak(extra.get("prisoners") or {})
        rows.append({
            "title": props.get("title"),
            "category": props.get("category"),
            "activity": _activity(props.get("summary")),
            "place_text": props.get("place_text"),
            "region": props.get("region"),
            "district": props.get("district"),
            "camp_region": extra.get("camp_region"),
            "period_raw": props.get("period_raw"),
            "year_from": props.get("year_from"),
            "year_to": props.get("year_to"),
            "date_approx": "да" if props.get("date_approx") else "нет",
            "peak": peak,
            "peak_year": peak_year,
            "lat": lat,
            "lon": lon,
            "url": props.get("url"),
            "published": "нет" if props.get("confidence") == "unpublished_source" else "да",
            "summary": props.get("summary"),
            "camp_id": _camp_id(props.get("source_id")),
            "uid": props.get("uid"),
            "source_id": props.get("source_id"),
            "prisoners": prisoners,
        })
    # Человек ищет лагерь по названию, а потом смотрит, где тот стоял в
    # нужный ему год.
    # Место одного управления держим подряд: колонка «Первая запись
    # управления» опирается на соседство строк, а не на пересчёт по всему
    # столбцу — иначе сводка считала бы полтора миллиона ячеек.
    rows.sort(key=lambda r: ((r["title"] or "").lower(), r["camp_id"] or 0,
                             r["year_from"] or 0))
    return rows


def _activity(summary: str | None) -> str | None:
    """Вид работ отдельным полем слой не хранит — вынимаем из описания."""
    for part in (summary or "").split("; "):
        found = _RE_ACTIVITY.match(part.strip().rstrip("."))
        if found:
            return found.group(1)
    return None


def _peak(prisoners: dict) -> tuple[int | None, int | None]:
    """Наибольшая учтённая численность и её год; нули за численность не считаются."""
    best, best_year = None, None
    for year, count in prisoners.items():
        if not isinstance(count, int) or count <= 0 or not str(year).isdigit():
            continue
        if best is None or count > best:
            best, best_year = count, int(year)
    return best, best_year


def _camp_id(source_id: str | None):
    """`source_id` — «управление:место»; управление нужно отдельной колонкой."""
    head = str(source_id or "").split(":")[0]
    return int(head) if head.isdigit() else None


# --- запись книги -----------------------------------------------------------

def build(rows: list[dict], out: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    # Значения формул openpyxl не считает: без этой пометки сводка открывается
    # пустой, пока читатель не нажмёт пересчёт руками.
    wb.calculation.fullCalcOnLoad = True
    wb.remove(wb.active)
    camps = wb.create_sheet("Лагеря")
    stats = wb.create_sheet("Численность по годам")
    total = wb.create_sheet("Сводка")
    about = wb.create_sheet("Об источнике")

    sheet_camps(camps, rows)
    sheet_stats(stats, rows)
    sheet_total(total, rows)
    sheet_about(about)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _head(ws, headers: list[str], *, height: int = 32) -> None:
    """Шапка листа: белым по тёмному, с переносом, закреплена."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="44546A")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = height


def _published_peak(rows: list[dict]) -> int:
    """Наибольшее число из карточек, которые проект у себя показывает."""
    published = [r["peak"] for r in rows if r.get("peak") and r.get("published") == "да"]
    return max(published) if published else 0


def _suspicious(row: dict, ceiling: int) -> bool:
    """Число выше всех опубликованных, да ещё в неопубликованной карточке.

    В ответе API попадается такая запись — «Новый лагерь», 456 789 человек:
    название общее, цифры идут подряд, карточку проект не показывает. Похоже
    на пробную запись, но доказательства этому у нас нет, а выбрасывать
    молча репозиторий не разрешает. Поэтому не отбор, а пометка — и проверка
    общая, а не по названию: заведут такую же ещё раз, пометится и она.
    """
    return bool(row.get("peak") and row.get("published") == "нет" and row["peak"] > ceiling)


def sheet_camps(ws, rows: list[dict]) -> None:
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    keys = [k for k, _, _ in CAMP_COLUMNS]
    _head(ws, [title for _, title, _ in CAMP_COLUMNS])
    for i, (_, _, width) in enumerate(CAMP_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    col = {k: get_column_letter(i) for i, (k, _, _) in enumerate(CAMP_COLUMNS, start=1)}
    body = Font(name=FONT, size=10)
    link = Font(name=FONT, size=10, color="0563C1", underline="single")
    ceiling = _published_peak(rows)

    for n, row in enumerate(rows, start=1):
        line = n + 1
        values = dict(row, no=n)
        # Полнота строки и принадлежность управлению считаются формулами:
        # книга остаётся живой, если её отсортируют или дополнят.
        values["years_known"] = f"=COUNT('Численность по годам'!D{line}:AT{line})"
        values["first_of_camp"] = (f"=IF({col['camp_id']}{line}={col['camp_id']}{line - 1},0,1)"
                                   if n > 1 else 1)
        ws.append([values.get(k) for k in keys])

        for cell in ws[line]:
            cell.font = body
            cell.alignment = Alignment(vertical="top")
        for key in ("peak", "camp_id", "peak_year", "year_from", "year_to"):
            ws[f"{col[key]}{line}"].number_format = "#,##0" if key == "peak" else "0"
        for key in ("lat", "lon"):
            ws[f"{col[key]}{line}"].number_format = "0.0000"
        url = ws[f"{col['url']}{line}"]
        if row.get("url"):
            url.hyperlink = row["url"]
            url.font = link
        if _suspicious(row, ceiling):
            # Правило репозитория: спорное помечается и остаётся в данных.
            cell = ws[f"{col['peak']}{line}"]
            cell.comment = Comment(SUSPICIOUS_NOTE, "выгрузка gulag_camps", height=120, width=340)
            cell.font = Font(name=FONT, size=10, color="C00000")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{col['source_id']}{ws.max_row}"


def sheet_stats(ws, rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    years = list(range(YEAR_MIN, YEAR_MAX + 1))
    _head(ws, ["№", "Название", "Место (как в источнике)"]
              + [str(y) for y in years] + ["Лет с числами", "Наибольшее"])
    for letter, width in (("A", 6), ("B", 44), ("C", 52)):
        ws.column_dimensions[letter].width = width
    first_year_col, last_year_col = "D", get_column_letter(3 + len(years))
    for i in range(4, 4 + len(years)):
        ws.column_dimensions[get_column_letter(i)].width = 8
    known_col = get_column_letter(4 + len(years))
    max_col = get_column_letter(5 + len(years))
    ws.column_dimensions[known_col].width = 12
    ws.column_dimensions[max_col].width = 12

    body = Font(name=FONT, size=10)
    for n, row in enumerate(rows, start=1):
        line = n + 1
        span = f"{first_year_col}{line}:{last_year_col}{line}"
        # Пусто — числа нет; ноль — в отчёте стоял ноль. Разница смысловая,
        # поэтому None и 0 не сводятся к одному виду.
        counts = [row["prisoners"].get(y) if isinstance(row["prisoners"].get(y), int) else None
                  for y in years]
        ws.append([n, row["title"], row["place_text"]] + counts
                  + [f"=COUNT({span})", f'=IF(COUNT({span})=0,"",MAX({span}))'])
        for cell in ws[line]:
            cell.font = body
            cell.alignment = Alignment(vertical="top")
        for i in range(4, 6 + len(years)):
            ws.cell(row=line, column=i).number_format = "#,##0"

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{max_col}{ws.max_row}"


def sheet_total(ws, rows: list[dict]) -> None:
    """Сводка. Все числа — формулы: книга остаётся живой после правок листов."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    last = len(rows) + 1
    src = f"Лагеря!$C$2:$C${last}"           # тип лагеря
    ref = {
        "category": f"Лагеря!$C$2:$C${last}",
        "activity": f"Лагеря!$D$2:$D${last}",
        "camp_region": f"Лагеря!$H$2:$H${last}",
    }
    camps_col = f"Лагеря!$V$2:$V${last}"     # 1 у первой записи управления
    peak_col = f"Лагеря!$M$2:$M${last}"      # наибольшая численность

    title_font = Font(name=FONT, bold=True, size=13)
    head_font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="44546A")
    body = Font(name=FONT, size=10)
    bold = Font(name=FONT, bold=True, size=10)
    note = Font(name=FONT, italic=True, size=9, color="595959")
    dash = '#,##0;-#,##0;"—"'

    for letter, width in (("A", 46), ("B", 13), ("C", 13), ("D", 13), ("E", 16)):
        ws.column_dimensions[letter].width = width

    line = 1

    def put(values, *, fonts=None, formats=None, height=None):
        nonlocal line
        ws.append(values)
        for i, cell in enumerate(ws[line]):
            cell.font = (fonts or body) if not isinstance(fonts, list) else fonts[i]
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if i == 0 else "right")
            if formats and i < len(formats) and formats[i]:
                cell.number_format = formats[i]
        if height:
            ws.row_dimensions[line].height = height
        line += 1
        return line - 1

    def section(text, comment=None):
        put([text], fonts=title_font, height=26)
        if comment:
            put([comment], fonts=note, height=14)

    def header(cells):
        row = put(cells, fonts=[head_font] * len(cells))
        for cell in ws[row]:
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", horizontal="center",
                                       wrap_text=True)
        ws.row_dimensions[row].height = 30

    section("Слой в целом")
    for label, formula, fmt in [
        ("Записей — мест размещения управлений", f"=COUNTA(Лагеря!$B$2:$B${last})", "#,##0"),
        ("Лагерных управлений", f"=SUM({camps_col})", "#,##0"),
        ("Карточек опубликовано проектом", f'=COUNTIF(Лагеря!$S$2:$S${last},"да")', "#,##0"),
        ("Карточек проект у себя не показывает", f'=COUNTIF(Лагеря!$S$2:$S${last},"нет")', "#,##0"),
        ("Записей с учтённой численностью", f"=COUNT({peak_col})", "#,##0"),
        ("Наибольшая численность в одной записи", f"=MAX({peak_col})", "#,##0"),
        ("Самый ранний год работы", f"=MIN(Лагеря!$J$2:$J${last})", "0"),
        ("Самый поздний год работы", f"=MAX(Лагеря!$K$2:$K${last})", "0"),
    ]:
        put([label, formula], formats=[None, fmt])
    put([])

    def group(caption, comment, field, values):
        section(caption, comment)
        header(["Значение", "Мест", "Управлений", "С числами", "Наибольшая численность"])
        first = line
        rng = ref[field]
        for value in values:
            if value is None:
                put(["(не указано)",
                     f'=SUMPRODUCT(({rng}="")*1)',
                     f'=SUMPRODUCT(({rng}="")*{camps_col})',
                     f'=SUMPRODUCT(({rng}="")*({peak_col}>0))',
                     f'=SUMPRODUCT(MAX(({rng}="")*{peak_col}))'],
                    formats=[None, "#,##0", "#,##0", "#,##0", dash])
            else:
                r = line
                put([value,
                     f"=COUNTIF({rng},$A{r})",
                     f"=SUMIFS({camps_col},{rng},$A{r})",
                     f'=COUNTIFS({rng},$A{r},{peak_col},">0")',
                     f"=SUMPRODUCT(MAX(({rng}=$A{r})*{peak_col}))"],
                    formats=[None, "#,##0", "#,##0", "#,##0", dash])
        totals = put(["Итого"] + [f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{line - 1})"
                                  for c in range(2, 5)] + [""],
                     fonts=bold, formats=[None, "#,##0", "#,##0", "#,##0", None])
        ws[f"A{totals}"].font = bold
        put([])

    group("По типу лагеря",
          "Справочник проекта. «Итого» сходится с числом записей слоя.",
          "category", _values(rows, "category"))
    group("По региону — группировка «Карты ГУЛАГа»",
          "Шестнадцать крупных регионов, которыми проект делит карту.",
          "camp_region", _values(rows, "camp_region"))
    group("По виду работ",
          "Производственный профиль из справочника проекта; написание — как в источнике.",
          "activity", _values(rows, "activity"))

    section("Учтённая численность по годам",
            "Сумма того, что учтено в этих карточках, а не численность системы "
            "лагерей в этом году: у большинства мест числа есть лишь за часть лет.")
    header(["Год", "Записей с числом", "Из них больше нуля",
            "Сумма учтённого", "Наибольшее в одной записи"])
    for i, year in enumerate(range(YEAR_MIN, YEAR_MAX + 1)):
        column = get_column_letter(4 + i)
        span = f"'Численность по годам'!{column}$2:{column}${last}"
        put([year, f"=COUNT({span})", f'=COUNTIF({span},">0")',
             f"=SUM({span})", f"=MAX({span})"],
            formats=["0", "#,##0", "#,##0", "#,##0", dash])

    ws.freeze_panes = "A2"


def _values(rows: list[dict], field: str) -> list:
    """Значения поля по убыванию частоты; пустое, если есть, — последней строкой."""
    counts: dict = {}
    blanks = 0
    for row in rows:
        value = row.get(field)
        if value:
            counts[value] = counts.get(value, 0) + 1
        else:
            blanks += 1
    ordered = sorted(counts, key=lambda v: (-counts[v], v))
    return ordered + ([None] if blanks else [])


def sheet_about(ws) -> None:
    from openpyxl.styles import Alignment, Font

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 104
    line = 1
    for block in ABOUT:
        kind, rest = block[0], block[1:]
        if kind == "h":
            ws.cell(row=line, column=1, value=rest[0]).font = Font(name=FONT, bold=True, size=13)
            ws.row_dimensions[line].height = 26
        elif kind == "p":
            cell = ws.cell(row=line, column=1, value=rest[0])
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=line, start_column=1, end_row=line, end_column=2)
            ws.row_dimensions[line].height = 14 * (1 + len(rest[0]) // 130)
        else:
            key = ws.cell(row=line, column=1, value=rest[0])
            key.font = Font(name=FONT, bold=True, size=10)
            key.alignment = Alignment(vertical="top")
            val = ws.cell(row=line, column=2, value=rest[1])
            val.font = Font(name=FONT, size=10)
            val.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[line].height = 14 * (1 + len(rest[1]) // 100)
        line += 1

    for title, text in COLUMN_NOTES.items():
        key = ws.cell(row=line, column=1, value=title)
        key.font = Font(name=FONT, bold=True, size=10)
        key.alignment = Alignment(vertical="top")
        val = ws.cell(row=line, column=2, value=text)
        val.font = Font(name=FONT, size=10)
        val.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[line].height = 14 * (1 + len(text) // 100)
        line += 1


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--src", type=Path, default=DEFAULT_SRC,
                        help=f"слой в GeoJSON (по умолчанию {DEFAULT_SRC.relative_to(ROOT)})")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT,
                        help=f"куда писать книгу (по умолчанию {DEFAULT_OUT.relative_to(ROOT)})")
    args = parser.parse_args(argv)

    if not args.src.exists():
        print(f"Нет файла {args.src}: соберите слой — "
              f"python3 scripts/harvest_gulag.py --build", file=sys.stderr)
        return 1

    rows = load(args.src)
    path = build(rows, args.out)
    camps = sum(1 for r in rows if r.get("camp_id"))
    unique = len({r["camp_id"] for r in rows if r.get("camp_id")})
    with_numbers = sum(1 for r in rows if r.get("peak"))
    print(f"Записано {path}")
    print(f"  мест размещения: {len(rows)}")
    print(f"  лагерных управлений: {unique} (из {camps} записей с идентификатором)")
    print(f"  записей с учтённой численностью: {with_numbers}")
    print("Формулы книга хранит без значений — пересчитайте её в LibreOffice "
          "или Excel, если читаете сводку программой.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
