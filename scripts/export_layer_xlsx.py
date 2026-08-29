#!/usr/bin/env python3
"""Выгрузка любого слоя единой схемы в XLSX — для чтения глазами.

    python3 scripts/export_layer_xlsx.py --layer repressions
    python3 scripts/export_layer_xlsx.py --layer disasters --out путь/файл.xlsx

Чем отличается от `write_xlsx` в `io_formats.py`. Тот пишет плоскую таблицу
всех полей схемы — годится для машины, но не для человека: колонки, у которых
на весь слой одно значение, повторяются в каждой строке, а качество слоя по
такой таблице не видно вовсе. Здесь три листа: записи, сводка на формулах и
лист об источнике, где оговорки не написаны от руки, а **посчитаны по самим
данным** — сколько записей без даты, сколько вне рамки проекта, сколько без
губернии. Оценка по факту, а не прикидка.

Слою лагерей ГУЛАГа выгрузка своя (`scripts/export_gulag_xlsx.py`): у него
в `extra` лежит численность заключённых по годам, и она разворачивается
отдельным листом, чего общая схема не умеет.

Ничего не выбрасывается. Записи вне рамки проекта и без датировки остаются
в таблице и помечаются колонкой — правило репозитория.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from histctx.geo import in_bbox  # noqa: E402
from histctx.registry import ALL_LAYERS  # noqa: E402
from histctx.schema import COLUMNS, COLUMNS_RU  # noqa: E402
from histctx.sources.wikidata import OPEN_END_YEAR  # noqa: E402
from histctx.xlsx_style import (  # noqa: E402
    Cursor, cell_value, head_row, link_font, new_workbook, warn_font,
)

GEOJSON_DIR = ROOT / "data" / "out" / "geojson"
QUERY_DIR = ROOT / "queries"
OUT_DIR = ROOT / "data" / "out" / "xlsx"

# Колонки, у которых на весь слой обычно одно значение: повторять их в каждой
# строке незачем, они уезжают на лист «Об источнике».
CONSTANT_CANDIDATES = ("layer", "layer_title", "group", "scope", "source", "license")

# Ширины по смыслу колонки; чего нет в списке — по умолчанию.
WIDTHS = {
    "title": 44, "summary": 62, "quote": 62, "url": 34, "image_url": 30,
    "place_text": 34, "region": 22, "regions": 30, "district": 20,
    "category": 30, "period_raw": 16, "uid": 26, "source_id": 14,
    "actor": 24, "work": 28, "lat": 11, "lon": 11,
}

# Вычисляемые колонки — их в схеме нет, они меряют качество записи.
COMPUTED = [
    ("in_frame", "В рамке проекта", 12),
    ("has_date", "Датировка есть", 12),
    ("in_period", "В периоде проекта", 13),
]


def load(slug: str, src: Path) -> tuple[list[dict], dict]:
    """Читает слой и возвращает записи вместе с шапкой файла."""
    with src.open(encoding="utf-8") as fh:
        payload = json.load(fh)
    features = payload.get("features") or []
    if not features:
        raise SystemExit(f"{src}: в слое нет записей — сначала соберите его")

    rows = []
    for feature in features:
        props = dict(feature.get("properties") or {})
        geometry = feature.get("geometry") or {}
        lat = lon = None
        if geometry.get("type") == "Point":
            coords = geometry.get("coordinates") or []
            if len(coords) >= 2:
                lon, lat = float(coords[0]), float(coords[1])
        props["lat"], props["lon"] = lat, lon
        props["in_frame"] = ("да" if (lat is not None and in_bbox(lat, lon))
                             else "нет" if lat is not None else "нет координаты")
        props["has_date"] = "да" if props.get("year_from") else "нет"
        # Верхняя граница интересующего периода объявлена в сборщике Викиданных
        # (OPEN_END_YEAR). Всё, что началось позже, к задаче проекта отношения
        # не имеет — но выбрасывать это не наше дело, помечаем.
        year = props.get("year_from")
        props["in_period"] = ("нет даты" if not year
                              else "да" if year <= OPEN_END_YEAR else "нет")
        rows.append(props)

    rows.sort(key=lambda r: ((r.get("title") or "").lower(), r.get("year_from") or 0))
    return rows, {k: v for k, v in payload.items() if k != "features"}


def split_columns(rows: list[dict]) -> tuple[list[str], dict]:
    """Делит поля на колонки таблицы и на постоянные — те уходят в описание."""
    present, constant = [], {}
    for name in COLUMNS:
        values = {r.get(name) for r in rows if r.get(name) not in (None, "")}
        if not values:
            continue                       # поле пустое у всего слоя — не колонка
        if name in CONSTANT_CANDIDATES and len(values) == 1:
            constant[COLUMNS_RU.get(name, name)] = values.pop()
            continue
        present.append(name)
    return present, constant


def measure(rows: list[dict], columns: list[str]) -> list[tuple[str, str]]:
    """Оговорки о слое, посчитанные по данным, а не написанные от руки."""
    total = len(rows)
    years = [r["year_from"] for r in rows if r.get("year_from")]
    out = [("Записей в слое", f"{total:,}".replace(",", " "))]

    def share(n: int, what: str) -> str:
        return f"{n} из {total} ({n * 100 // total}%) — {what}"

    no_date = sum(1 for r in rows if r["has_date"] == "нет")
    if no_date:
        out.append(("Записей без датировки",
                    share(no_date, "год не проставлен ни один. На ленте времени "
                                   "такая запись не встанет, на карте — встанет")))
    if years:
        out.append(("Годы, где датировка есть", f"{min(years)}–{max(years)}"))

    later = sum(1 for r in rows if r["in_period"] == "нет")
    if later:
        out.append((f"Записей позже верхней границы проекта ({OPEN_END_YEAR})",
                    share(later, "объект возник после периода, ради которого "
                                 "слой собирается: современные тюрьмы и колонии. "
                                 "Помечены колонкой «В периоде проекта»")))

    outside = sum(1 for r in rows if r["in_frame"] == "нет")
    if outside:
        out.append(("Записей вне рамки проекта",
                    share(outside, "точка вне охвата Российской империи и СССР "
                                   "(geo.BBOX_RU). Не выброшены: помечены колонкой "
                                   "«В рамке проекта»")))
    if "region" in columns:
        no_region = sum(1 for r in rows if not r.get("region"))
        if no_region:
            out.append(("Записей без губернии или области",
                        share(no_region, "разбор места не дал привязки")))
    if "category" in columns:
        cats = {r.get("category") for r in rows if r.get("category")}
        out.append(("Разных категорий", f"{len(cats)} — разбор по ним на листе «Сводка»"))
    return out


# --- запись книги -----------------------------------------------------------

def build(rows: list[dict], columns: list[str], constant: dict,
          spec, header: dict, out: Path) -> Path:
    wb = new_workbook()

    keys = columns + [k for k, _, _ in COMPUTED]
    titles = ([COLUMNS_RU.get(c, c) for c in columns]
              + [t for _, t, _ in COMPUTED])
    widths = ([WIDTHS.get(c, 15) for c in columns] + [w for _, _, w in COMPUTED])

    sheet_rows(wb.create_sheet("Записи"), rows, keys, titles, widths)
    sheet_total(wb.create_sheet("Сводка"), rows, keys, len(rows))
    sheet_about(wb.create_sheet("Об источнике"), rows, columns, constant, spec, header)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def sheet_rows(ws, rows, keys, titles, widths) -> None:
    from openpyxl.utils import get_column_letter

    head_row(ws, ["№"] + titles)
    ws.column_dimensions["A"].width = 6
    for i, width in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = width

    link, warn = link_font(), warn_font()
    url_cols = [i for i, k in enumerate(keys, start=2) if k in ("url", "image_url")]
    number_cols = [(keys.index(k) + 2, f) for k, f in
                   (("lat", "0.0000"), ("lon", "0.0000"),
                    ("year_from", "0"), ("year_to", "0")) if k in keys]
    mark_cols = [(keys.index(k) + 2, k) for k in ("in_frame", "in_period")]

    for n, row in enumerate(rows, start=1):
        line = n + 1
        ws.append([n] + [cell_value(row.get(k)) for k in keys])
        for i in url_cols:
            cell = ws.cell(row=line, column=i)
            if cell.value:
                cell.hyperlink, cell.font = cell.value, link
        for i, fmt in number_cols:
            ws.cell(row=line, column=i).number_format = fmt
        # Запись вне рамки или вне периода видно сразу, а не после сортировки.
        for i, key in mark_cols:
            if row[key] != "да":
                ws.cell(row=line, column=i).font = warn

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keys) + 1)}{ws.max_row}"


def sheet_total(ws, rows: list[dict], keys: list[str], count: int) -> None:
    """Сводка. Числа — формулы: книга остаётся живой после правок и сортировок."""
    from openpyxl.utils import get_column_letter

    last = count + 1

    def rng(key):
        letter = get_column_letter(keys.index(key) + 2)
        return f"Записи!${letter}$2:${letter}${last}"

    for letter, width in (("A", 54), ("B", 13), ("C", 13), ("D", 13)):
        ws.column_dimensions[letter].width = width
    cur = Cursor(ws)

    cur.section("Слой в целом")
    stats = [
        ("Записей", f"=COUNTA({rng('title')})", "#,##0"),
        ("С датировкой", f'=COUNTIF({rng("has_date")},"да")', "#,##0"),
        ("Без датировки", f'=COUNTIF({rng("has_date")},"нет")', "#,##0"),
        ("В рамке проекта (Российская империя и СССР)",
         f'=COUNTIF({rng("in_frame")},"да")', "#,##0"),
        ("Вне рамки проекта", f'=COUNTIF({rng("in_frame")},"нет")', "#,##0"),
        (f"Позже верхней границы проекта ({OPEN_END_YEAR})",
         f'=COUNTIF({rng("in_period")},"нет")', "#,##0"),
    ]
    if "year_from" in keys:
        stats.append(("Самый ранний год с датировкой", f"=MIN({rng('year_from')})", "0"))
        end = "year_to" if "year_to" in keys else "year_from"
        stats.append(("Самый поздний год с датировкой", f"=MAX({rng(end)})", "0"))
    for label, formula, fmt in stats:
        cur.row([label, formula], formats=[None, fmt])
    cur.blank()

    if "category" in keys:
        group_block(cur, rows, "category", rng, "По категориям",
                    "Категория берётся из источника как есть и не правится.")
    if "region" in keys:
        group_block(cur, rows, "region", rng, "По губерниям и областям",
                    "Показаны те, где больше одной записи; одиночные сведены "
                    "в отдельную строку, чтобы «Итого» сходилось.", min_count=2)
    if "year_from" in keys:
        decade_block(cur, rows, rng)

    ws.freeze_panes = "A2"


def group_block(cur: Cursor, rows: list[dict], field: str, rng, caption: str,
                comment: str, *, min_count: int = 1) -> None:
    """Разбор по значению поля: записей, из них в рамке и с датировкой."""
    import collections

    counts = collections.Counter(r.get(field) for r in rows if r.get(field))
    blanks = sum(1 for r in rows if not r.get(field))
    shown = [v for v, n in counts.most_common() if n >= min_count]
    hidden = [v for v, n in counts.most_common() if n < min_count]
    value_rng, frame_rng, date_rng = rng(field), rng("in_frame"), rng("has_date")

    cur.section(caption, comment)
    cur.header(["Значение", "Записей", "В рамке", "С датой"])
    first = last = None
    for value in shown:
        line = cur.row([value, None, None, None],
                       formats=[None, "#,##0", "#,##0", "#,##0"])
        cur.ws[f"B{line}"] = f"=COUNTIF({value_rng},$A{line})"
        cur.ws[f"C{line}"] = f'=COUNTIFS({value_rng},$A{line},{frame_rng},"да")'
        cur.ws[f"D{line}"] = f'=COUNTIFS({value_rng},$A{line},{date_rng},"да")'
        first = first or line
        last = line
    if hidden:
        # Не «прочие», а именно значения с одной записью. Число считается
        # вычитанием, чтобы «Итого» сошлось с числом записей слоя.
        line = cur.row([f"по одной записи — {len(hidden)} значений",
                        None, None, None],
                       formats=[None, "#,##0", "#,##0", "#,##0"])
        rest = f"-SUM(B{first}:B{last})" if first else ""
        cur.ws[f"B{line}"] = f'=SUMPRODUCT(({value_rng}<>"")*1){rest}'
        cur.ws[f"C{line}"] = (f'=SUMPRODUCT(({value_rng}<>"")*({frame_rng}="да"))'
                              + (f"-SUM(C{first}:C{last})" if first else ""))
        cur.ws[f"D{line}"] = (f'=SUMPRODUCT(({value_rng}<>"")*({date_rng}="да"))'
                              + (f"-SUM(D{first}:D{last})" if first else ""))
        first = first or line
        last = line
    if blanks:
        line = cur.row(["(не указано)", None, None, None],
                       formats=[None, "#,##0", "#,##0", "#,##0"])
        cur.ws[f"B{line}"] = f'=SUMPRODUCT(({value_rng}="")*1)'
        cur.ws[f"C{line}"] = f'=SUMPRODUCT(({value_rng}="")*({frame_rng}="да"))'
        cur.ws[f"D{line}"] = f'=SUMPRODUCT(({value_rng}="")*({date_rng}="да"))'
        first = first or line
        last = line
    total = cur.row(["Итого", None, None, None], font=cur.bold,
                    formats=[None, "#,##0", "#,##0", "#,##0"])
    for letter in "BCD":
        cur.ws[f"{letter}{total}"] = f"=SUM({letter}{first}:{letter}{last})"
    cur.blank()


def decade_block(cur: Cursor, rows: list[dict], rng) -> None:
    """Раскладка по десятилетиям — только там, где датировка есть."""
    years = [r["year_from"] for r in rows if r.get("year_from")]
    if not years:
        return
    year_rng = rng("year_from")
    cur.section("По десятилетиям",
                "Только записи с датировкой, по году начала. Записи без года "
                "сюда не попадают — их число на строке «Без датировки» выше.")
    cur.header(["Десятилетие", "Записей", "", ""])
    first = last = None
    for decade in range((min(years) // 10) * 10, (max(years) // 10) * 10 + 10, 10):
        line = cur.row([f"{decade}-е", None, None, None],
                       formats=[None, "#,##0", None, None])
        cur.ws[f"B{line}"] = (f'=COUNTIFS({year_rng},">={decade}",'
                              f'{year_rng},"<={decade + 9}")')
        first = first or line
        last = line
    total = cur.row(["Итого записей с датировкой", None, None, None],
                    font=cur.bold, formats=[None, "#,##0", None, None])
    cur.ws[f"B{total}"] = f"=SUM(B{first}:B{last})"
    cur.blank()


def query_classes(slug: str) -> list[str]:
    """Классы, по которым отбирался слой, — из заголовка SPARQL-запроса.

    Читателю слоя важно понимать не только что в нём есть, но и по какому
    признаку это туда попало: «место массового убийства» ловит и расстрельный
    полигон под Москвой, и лагерь смерти в оккупированной Польше.
    """
    path = QUERY_DIR / f"{slug}.rq"
    if not slug or not path.exists():
        return []
    out = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line.startswith("#"):
            break                      # шапка кончилась, дальше сам запрос
        if line.startswith("# @qid "):
            parts = line[len("# @qid "):].split(maxsplit=1)
            qid = parts[0]
            name = parts[1] if len(parts) > 1 else ""
            out.append(f"{name} ({qid})" if name else qid)
    return out


def sheet_about(ws, rows: list[dict], columns: list[str], constant: dict,
                spec, header: dict) -> None:
    """Откуда слой, что в нём одинаково у всех записей и чего в нём ждать."""
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 100
    cur = Cursor(ws)

    def pair(key, value):
        cur.pair(key, value, per_line=95)

    cur.section(header.get("name") or (spec.title if spec else "Слой"))
    cur.row(["Выгрузка слоя из репозитория исторического контекста. "
             "Пересобирается командой "
             "`python3 scripts/export_layer_xlsx.py --layer <слой>`."],
            font=cur.note, height=14)
    cur.blank()

    cur.section("Откуда данные")
    if spec is not None:
        pair("Слой", f"{spec.slug} — {spec.title}")
        if spec.description:
            pair("Что в нём", spec.description)
        if spec.url:
            pair("Проект-источник", spec.url)
    said = {COLUMNS_RU[k] for k in ("layer", "layer_title", "group")} if spec else set()
    for key, value in constant.items():
        if key not in said:
            pair(key, value)
    classes = query_classes(spec.slug if spec else "")
    if classes:
        pair("Отобрано по классам источника",
             "; ".join(classes) + ". В слой попадает всё, что относится к этим "
             "классам и стоит в рамке проекта, — в том числе то, к советским "
             "репрессиям отношения не имеющее. Разбор по категориям, как их "
             "называет источник, — на листе «Сводка»")
    cur.blank()

    cur.section("Что нужно знать, прежде чем считать",
                "Оговорки не написаны от руки — они посчитаны по этим самым данным.")
    for key, value in measure(rows, columns):
        pair(key, value)
    cur.blank()

    cur.section("Колонки")
    pair("№", "Порядковый номер строки в этой выгрузке, не идентификатор записи")
    for name in columns:
        pair(COLUMNS_RU.get(name, name), COLUMN_NOTES.get(name, "Поле единой схемы"))
    pair("В рамке проекта", "Попадает ли точка в охват Российской империи и СССР "
                            "(`geo.BBOX_RU`). «нет» — запись оставлена и помечена, "
                            "а не выброшена")
    pair("Датировка есть", "Проставлен ли год начала. «нет» — запись есть на карте, "
                           "но не встанет на ленту времени")
    pair("В периоде проекта", f"Начался ли объект не позже {OPEN_END_YEAR} года — "
                              "верхней границы, объявленной в сборщике "
                              "(`histctx.sources.wikidata.OPEN_END_YEAR`). «нет» — "
                              "запись оставлена и помечена, а не выброшена")


COLUMN_NOTES = {
    "uid": "Идентификатор записи в слое репозитория",
    "title": "Название объекта из источника",
    "category": "Класс объекта из источника, как он там записан",
    "lat": "Широта, градусы, WGS 84",
    "lon": "Долгота, градусы, WGS 84",
    "place_text": "Текст привязки к месту из источника",
    "region": "Губерния или область, разобранная из текста места",
    "regions": "Затронутые губернии — для записей без одной точки",
    "district": "Уезд или район, разобранный из текста места",
    "year_from": "Год начала. Пусто — датировки в источнике нет",
    "year_to": "Год конца. У открытого срока подтянут к верхней границе охвата",
    "date_precision": "Насколько точна дата: year — год, part — период, "
                      "open — срок открыт, unknown — датировки нет",
    "date_approx": "«да» — точная дата неизвестна: взят период или открытый срок",
    "period_raw": "Период, как он записан в слое",
    "actor": "Действующее лицо или автор",
    "work": "Произведение или событие",
    "summary": "Короткое описание из источника",
    "quote": "Цитата из источника",
    "url": "Ссылка на карточку источника",
    "image_url": "Ссылка на изображение. Сам файл не копируется — только адрес",
    "source_id": "Идентификатор записи в источнике",
    "confidence": "Насколько записи можно верить по мерке сборщика",
}


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--layer", required=True, help="slug слоя, например repressions")
    parser.add_argument("--src", type=Path, help="файл слоя (по умолчанию из data/out/geojson)")
    parser.add_argument("--out", type=Path, help="куда писать книгу")
    args = parser.parse_args(argv)

    src = args.src or GEOJSON_DIR / f"{args.layer}.geojson"
    out = args.out or OUT_DIR / f"{args.layer}.xlsx"
    if not src.exists():
        print(f"Нет файла {src}. Собранные слои: "
              f"{', '.join(sorted(p.stem for p in GEOJSON_DIR.glob('*.geojson')))}",
              file=sys.stderr)
        return 1

    spec = next((s for s in ALL_LAYERS if s.slug == args.layer), None)
    rows, header = load(args.layer, src)
    columns, constant = split_columns(rows)
    path = build(rows, columns, constant, spec, header, out)

    print(f"Записано {path}")
    print(f"  записей: {len(rows)}")
    for key, value in measure(rows, columns):
        if key != "Записей в слое":
            print(f"  {key.lower()}: {value}")
    print("Формулы книга хранит без значений — пересчитайте её в LibreOffice "
          "или Excel, если читаете сводку программой.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
